from flask import Blueprint, render_template, request, jsonify, send_file, current_app
from flask_login import login_required, current_user
from database import db
from models.user import ScanHistory
from datetime import datetime, timedelta
import io, json

reports_bp = Blueprint('reports', __name__)


@reports_bp.route('/')
@login_required
def index():
    return render_template('reports/index.html')


@reports_bp.route('/summary')
@login_required
def summary():
    days = request.args.get('days', 30, type=int)
    since = datetime.utcnow() - timedelta(days=days)
    scans = ScanHistory.query.filter(
        ScanHistory.user_id == current_user.id,
        ScanHistory.created_at >= since
    ).all()
    report = _build_summary(scans, days)
    return render_template('reports/summary.html', report=report)


@reports_bp.route('/export/pdf')
@login_required
def export_pdf():
    days = request.args.get('days', 30, type=int)
    since = datetime.utcnow() - timedelta(days=days)
    scans = ScanHistory.query.filter(
        ScanHistory.user_id == current_user.id,
        ScanHistory.created_at >= since
    ).all()
    pdf_bytes = _generate_pdf(scans, current_user)
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'cybershield_report_{datetime.utcnow().strftime("%Y%m%d")}.pdf'
    )


@reports_bp.route('/export/excel')
@login_required
def export_excel():
    days = request.args.get('days', 30, type=int)
    since = datetime.utcnow() - timedelta(days=days)
    scans = ScanHistory.query.filter(
        ScanHistory.user_id == current_user.id,
        ScanHistory.created_at >= since
    ).all()
    excel_bytes = _generate_excel(scans)
    return send_file(
        io.BytesIO(excel_bytes),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'cybershield_report_{datetime.utcnow().strftime("%Y%m%d")}.xlsx'
    )


@reports_bp.route('/export/csv')
@login_required
def export_csv():
    days = request.args.get('days', 30, type=int)
    since = datetime.utcnow() - timedelta(days=days)
    scans = ScanHistory.query.filter(
        ScanHistory.user_id == current_user.id,
        ScanHistory.created_at >= since
    ).order_by(ScanHistory.created_at.desc()).all()

    import csv
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Type', 'Input', 'Result', 'Risk Score', 'Risk Level', 'Date'])
    for s in scans:
        writer.writerow([s.id, s.scan_type, s.input_data or '', s.result or '', s.risk_score, s.risk_level, s.created_at.strftime('%Y-%m-%d %H:%M')])

    return send_file(
        io.BytesIO(output.getvalue().encode()),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'cybershield_report_{datetime.utcnow().strftime("%Y%m%d")}.csv'
    )


def _build_summary(scans, days):
    by_type = {}
    by_risk = {'safe': 0, 'suspicious': 0, 'dangerous': 0}
    for s in scans:
        by_type[s.scan_type] = by_type.get(s.scan_type, 0) + 1
        if s.risk_level in by_risk:
            by_risk[s.risk_level] += 1
    avg_risk = round(sum(s.risk_score for s in scans) / len(scans), 1) if scans else 0
    return {
        'total_scans': len(scans),
        'by_type': by_type,
        'by_risk': by_risk,
        'avg_risk_score': avg_risk,
        'period_days': days,
        'generated_at': datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC'),
        'user': current_user.full_name or current_user.username,
        'recommendations': _get_recommendations(by_risk, by_type)
    }


def _get_recommendations(by_risk, by_type):
    recs = []
    if by_risk.get('dangerous', 0) > 0:
        recs.append('Multiple dangerous threats detected. Review and block flagged URLs/emails.')
    if by_type.get('password', 0) > 0:
        recs.append('Consider updating weak passwords identified in scans.')
    if by_type.get('malware', 0) > 0:
        recs.append('Malware scan activity detected. Ensure antivirus is up to date.')
    if not recs:
        recs.append('No critical issues found. Maintain security hygiene and keep scanning.')
    return recs


def _generate_pdf(scans, user):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph('CyberShield AI - Security Report', styles['Title']))
        elements.append(Paragraph(f'User: {user.full_name or user.username}', styles['Normal']))
        elements.append(Paragraph(f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}', styles['Normal']))
        elements.append(Spacer(1, 20))

        if scans:
            data = [['ID', 'Type', 'Input', 'Risk Level', 'Score', 'Date']]
            for s in scans[:50]:
                data.append([str(s.id), s.scan_type, (s.input_data or '')[:40], s.risk_level or '', str(s.risk_score), s.created_at.strftime('%Y-%m-%d')])

            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a1a2e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
            ]))
            elements.append(table)

        doc.build(elements)
        return buffer.getvalue()
    except ImportError:
        return b'PDF generation requires reportlab. Install with: pip install reportlab'


def _generate_excel(scans):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Scan History'

        headers = ['ID', 'Type', 'Input', 'Result', 'Risk Score', 'Risk Level', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(fill_type='solid', fgColor='1a1a2e')
            cell.alignment = Alignment(horizontal='center')

        for row, s in enumerate(scans, 2):
            ws.append([s.id, s.scan_type, s.input_data or '', s.result or '', s.risk_score, s.risk_level, s.created_at.strftime('%Y-%m-%d %H:%M')])
            if s.risk_level == 'dangerous':
                for col in range(1, 8):
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type='solid', fgColor='FFE0E0')

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except ImportError:
        return b''
